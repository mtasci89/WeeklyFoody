from __future__ import annotations

import json
import re
from pathlib import Path
from typing import Protocol

import yaml
from pydantic import BaseModel, Field


class IngredientInput(BaseModel):
    name: str
    quantity: float | None = None
    unit: str | None = None
    category: str | None = None
    note: str | None = None


class RecipeInput(BaseModel):
    name: str
    aliases: list[str] = Field(default_factory=list)
    cuisine: str | None = None
    category: str = "main"
    meal_type: str = "dinner"
    ingredients: list[IngredientInput] = Field(default_factory=list)
    servings: int = 4
    instructions: str | None = None
    prep_minutes: int | None = None
    cook_minutes: int | None = None
    tags: list[str] = Field(default_factory=list)
    protein_type: str | None = None
    vegetarian: bool = False
    seasonal: list[str] = Field(default_factory=list)
    source: str | None = None
    notes: str | None = None
    status: str = "approved"


class RecipeImporter(Protocol):
    def import_path(self, path: Path) -> list[RecipeInput]:
        ...


class StructuredRecipeImporter:
    def import_path(self, path: Path) -> list[RecipeInput]:
        records: list[dict] = []
        if path.is_dir():
            for child in sorted(path.iterdir()):
                if child.suffix.lower() in {".yaml", ".yml", ".json"}:
                    records.extend(self.import_path(child))
            return records  # type: ignore[return-value]

        raw = path.read_text(encoding="utf-8")
        if path.suffix.lower() == ".json":
            data = json.loads(raw)
        elif path.suffix.lower() in {".yaml", ".yml"}:
            data = yaml.safe_load(raw)
        else:
            raise ValueError(f"Unsupported recipe file: {path}")

        if isinstance(data, dict):
            data = [data]
        return [RecipeInput.model_validate(item) for item in data or []]


class ExcelMenuRecipeImporter:
    meal_labels = {
        "öğle yemeği": "lunch",
        "akşam": "dinner",
        "ara öğün": "snack",
    }
    day_names = {"pazartesi", "salı", "çarşamba", "perşembe", "cuma", "cumartesi", "pazar"}
    ignored_items = {
        "yoğurt",
        "salata",
        "yeşil salata",
        "1 dilim wasa fibre",
        "wasa fibre",
        "ayran",
    }

    def import_path(self, path: Path) -> list[RecipeInput]:
        from openpyxl import load_workbook

        workbook = load_workbook(path, data_only=True)
        recipes: dict[tuple[str, str], RecipeInput] = {}
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows(values_only=True):
                meal_type = self._row_meal_type(row)
                if not meal_type:
                    continue
                for value in row:
                    if not isinstance(value, str):
                        continue
                    for name in self._extract_names(value):
                        key = (name.casefold(), meal_type)
                        recipes.setdefault(
                            key,
                            RecipeInput(
                                name=name,
                                meal_type=meal_type,
                                category=self._category_for(name),
                                source=f"excel:{path.name}",
                                tags=["menu-import"],
                                notes="Imported from historical menu list; ingredients can be taught later.",
                            ),
                        )
        return sorted(recipes.values(), key=lambda recipe: (recipe.meal_type, recipe.name.casefold()))

    def _row_meal_type(self, row: tuple) -> str | None:
        for value in row:
            if isinstance(value, str):
                label = value.strip().casefold()
                if label in self.meal_labels:
                    return self.meal_labels[label]
        return None

    def _extract_names(self, value: str) -> list[str]:
        chunks = re.split(r"https?://\S+", value)
        names: list[str] = []
        for chunk in chunks:
            for line in chunk.splitlines():
                cleaned = self._clean_name(line)
                if cleaned:
                    names.append(cleaned)
        return names

    def _clean_name(self, raw: str) -> str | None:
        text = " ".join(raw.strip(" -\t,").split())
        if not text:
            return None
        lower = text.casefold()
        if lower in self.meal_labels or lower in self.day_names or lower in self.ignored_items:
            return None
        if text.startswith("(") and text.endswith(")"):
            return None
        text = re.sub(r"^\d+\s*(?:yemek kaşığı|bardak|gram|gr|g)\.?\s*", "", text, flags=re.IGNORECASE).strip()
        text = re.sub(r"\b1\s+dilim\s+wasa\s+fibre\b", "", text, flags=re.IGNORECASE).strip(" -,")
        text = re.sub(r"\s+", " ", text).strip()
        if not text or text.casefold() in self.ignored_items:
            return None
        if len(text) < 3:
            return None
        return text[:1].upper() + text[1:]

    def _category_for(self, name: str) -> str:
        return infer_recipe_category(name)


class PlainTextRecipeImporter:
    """Small pragmatic parser for Telegram recipe messages."""

    def parse_text(self, text: str) -> RecipeInput:
        source_url = None
        url_match = re.search(r"https?://\S+", text)
        if url_match:
            source_url = url_match.group(0).rstrip(".,)")
        cleaned = (
            text.replace("/addrecipe", "")
            .replace("Bu tarifi kaydet:", "")
            .replace("bu tarifi kaydet:", "")
            .replace("Kaydet:", "")
            .replace("kaydet:", "")
            .strip()
        )
        cleaned = re.sub(r"https?://\S+", "", cleaned).strip()
        lines = [line.strip(" -\t") for line in cleaned.splitlines() if line.strip()]
        if not lines:
            raise ValueError("Recipe text is empty")
        name = lines[0]
        ingredients: list[IngredientInput] = []
        instructions: list[str] = []
        ingredient_units = {"gr", "g", "gram", "kg", "ml", "l", "lt", "adet", "paket", "demet", "yemek kaşığı", "çay kaşığı"}
        for line in lines[1:]:
            lower = line.lower()
            if any(unit in lower.split() for unit in ingredient_units) or any(ch.isdigit() for ch in lower):
                parts = line.split()
                quantity = None
                unit = None
                name_parts = parts
                for idx, part in enumerate(parts):
                    normalized = part.replace(",", ".")
                    try:
                        quantity = float(normalized)
                        next_part = parts[idx + 1] if idx + 1 < len(parts) else None
                        if next_part and next_part.casefold() in ingredient_units:
                            unit = next_part
                            name_parts = parts[idx + 2 :]
                        else:
                            unit = "adet"
                            name_parts = parts[idx + 1 :]
                        break
                    except ValueError:
                        continue
                ingredient_name = " ".join(name_parts).strip() or line
                ingredients.append(IngredientInput(name=ingredient_name.lower(), quantity=quantity, unit=unit))
            else:
                instructions.append(line)
        return RecipeInput(
            name=name,
            category=infer_recipe_category(name),
            ingredients=ingredients,
            instructions="\n".join(instructions) or None,
            source=source_url or "telegram",
        )


def infer_recipe_category(name: str) -> str:
    lower = name.casefold()
    if any(token in lower for token in ("salata", "salatası")):
        return "salad"
    if any(token in lower for token in ("meze", "haydari", "humus", "cacık", "ezme", "muhammara", "şakşuka")):
        return "meze"
    if "çorba" in lower:
        return "soup"
    if any(token in lower for token in ("pilav", "makarna", "kinoa", "quinoa", "bulgur", "patates püresi")):
        return "side"
    return "main"
